import ttkbootstrap as tb
from ttkbootstrap.constants import *
from tkinter import Toplevel, Label, Entry, Button, messagebox, filedialog, StringVar, END, BOTH, LEFT, SOLID
from PIL import Image, ImageTk, UnidentifiedImageError
from db import obtener_productos, agregar_producto_db, actualizar_cantidad, eliminar_producto_db, actualizar_imagen
import os
import csv
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter

ASSETS_PATH = "assets/img"


def iniciar_app():
    app = tb.Window(themename="flatly")
    app.title("Inventario Pizzería")
    app.geometry("1000x700")

    topbar = tb.Frame(app, bootstyle="primary")
    topbar.pack(side=TOP, fill=X)

    titulo = tb.Label(topbar, text="Inventario Pizzería", font=("Segoe UI", 18, "bold"), foreground="white", bootstyle="inverse-primary")
    titulo.pack(side=LEFT, padx=20, pady=10)

    btn_eliminar = tb.Button(topbar, text="Eliminar Producto", bootstyle="danger", command=lambda: eliminar_producto_seleccionado())
    btn_eliminar.pack(side=RIGHT, padx=10)

    btn_agregar = tb.Button(topbar, text="Agregar Producto", bootstyle="success", command=lambda: agregar_producto())
    btn_agregar.pack(side=RIGHT, padx=10)

    frame_productos = tb.Frame(app)
    frame_productos.pack(fill=BOTH, expand=True, padx=10, pady=10)

    frame_filtros = tb.LabelFrame(app, text="Búsqueda y Filtros", bootstyle="info")
    frame_filtros.pack(fill=X, padx=10, pady=5)

    tb.Label(frame_filtros, text="Buscar por nombre:").pack(side=LEFT, padx=5)
    entry_buscar = tb.Entry(frame_filtros, width=35, foreground="gray")
    entry_buscar.pack(side=LEFT, padx=5)
    entry_buscar.insert(0, "Nombre del producto")

    tb.Label(frame_filtros, text="  Filtro:").pack(side=LEFT, padx=5)
    filtro_opcion = StringVar(value="Todos")
    combo_filtros = tb.Combobox(frame_filtros, textvariable=filtro_opcion, values=["Todos", "Poco stock", "Sin stock"], state="readonly", width=15)
    combo_filtros.pack(side=LEFT, padx=5)

    def exportar_excel():
        fecha_actual = datetime.now().strftime("%Y-%m-%d")
        ruta = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=f"informe_inventario_{fecha_actual}", filetypes=[("Archivo Excel", "*.xlsx")])
        if not ruta:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventario"

        encabezados = ["Estado", "Nombre", "Cantidad", "Última Modificación"]
        ws.append(encabezados)

        for item in inventario_total.get_children():
            fila = inventario_total.item(item, "values")
            ws.append(fila)

        for col in range(1, ws.max_column + 1):
            max_length = 0
            col_letter = get_column_letter(col)
            for cell in ws[col_letter]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

        wb.save(ruta)
        messagebox.showinfo("Éxito", "Inventario exportado correctamente.")

    btn_exportar = tb.Button(frame_filtros, text="Exportar a Excel", bootstyle="info", command=lambda: exportar_excel())
    btn_exportar.pack(side=RIGHT, padx=5)

    frame_listado = tb.LabelFrame(app, text="Lista de Productos", bootstyle="secondary")
    frame_listado.pack(fill=BOTH, expand=True, padx=10, pady=5)

    columnas = ("Estado", "Nombre", "Cantidad", "Última Modificación")
    inventario_total = tb.Treeview(frame_listado, columns=columnas, show="headings")

    for col in columnas:
        inventario_total.heading(col, text=col)

    inventario_total.column("Estado", width=150, anchor="center")
    inventario_total.column("Nombre", width=250, anchor="center")
    inventario_total.column("Cantidad", width=100, anchor="center")
    inventario_total.column("Última Modificación", width=200, anchor="center")

    inventario_total.pack(fill=BOTH, expand=True, padx=10, pady=5)

    def refrescar_tabla():
        for row in inventario_total.get_children():
            inventario_total.delete(row)

        productos = obtener_productos()
        filtro = filtro_opcion.get()
        buscar = entry_buscar.get()

        for nombre, cantidad, _, fecha_mod in productos:
            if buscar and buscar != "Nombre del producto" and buscar.lower() not in nombre.lower():
                continue

            estado = ""
            if cantidad == 0:
                estado = "Sin stock"
            elif cantidad < 5:
                estado = "Poco stock"

            if filtro == "Poco stock" and cantidad >= 5:
                continue
            if filtro == "Sin stock" and cantidad != 0:
                continue

            simbolo = ""
            if cantidad == 0:
                simbolo = "🔴 "
            elif cantidad < 5:
                simbolo = "🟡 "

            inventario_total.insert('', END, values=(f"{simbolo}{estado}", nombre, cantidad, fecha_mod))

    def refrescar_productos_superior():
        for widget in frame_productos.winfo_children():
            widget.destroy()

        productos = obtener_productos()

        if not productos:
            lbl_vacio = tb.Label(frame_productos, text="No hay productos cargados", font=("Segoe UI", 14), bootstyle="warning")
            lbl_vacio.pack(pady=20)
            return

        for idx, (nombre, cantidad, imagen, fecha_mod) in enumerate(productos):
            frame = tb.Frame(frame_productos, relief=SOLID, borderwidth=1, width=140, height=170)
            frame.grid_propagate(False)
            frame.grid(row=idx//6, column=idx%6, padx=5, pady=5)

            if imagen and os.path.exists(imagen):
                try:
                    img_pil = Image.open(imagen).resize((100, 100))
                    img = ImageTk.PhotoImage(img_pil)
                    lbl_img = Label(frame, image=img)
                    lbl_img.image = img
                    lbl_img.pack()
                except:
                    tb.Label(frame, text="Imagen inválida", bootstyle="danger").pack()

            tb.Label(frame, text=nombre, font=("Segoe UI", 11, "bold")).pack(pady=2)
            tb.Label(frame, text=f"Stock: {cantidad}", bootstyle="info").pack(pady=2)
            tb.Button(frame, text="Modificar", bootstyle="secondary", command=lambda n=nombre: abrir_modificar(n)).pack(pady=2)
            tb.Button(frame, text="Cambiar Imagen", bootstyle="info", command=lambda n=nombre: cambiar_imagen(n)).pack(pady=2)

    def abrir_modificar(nombre):
        top = Toplevel(app)
        top.title(f"Modificar {nombre}")
        top.geometry("300x150")
        top.transient(app)
        top.grab_set()

        productos = dict((p[0], p[1]) for p in obtener_productos())
        cantidad_actual = productos[nombre]
        cantidad_temp = tb.IntVar(value=cantidad_actual)

        tb.Label(top, text="Cantidad actual:").pack(pady=5)
        contenedor = tb.Frame(top)
        contenedor.pack(pady=5)

        tb.Button(contenedor, text="-", bootstyle="danger", command=lambda: cantidad_temp.set(max(0, cantidad_temp.get()-1))).pack(side=LEFT, padx=5)
        tb.Entry(contenedor, textvariable=cantidad_temp, width=10, justify="center").pack(side=LEFT)
        tb.Button(contenedor, text="+", bootstyle="success", command=lambda: cantidad_temp.set(cantidad_temp.get()+1)).pack(side=LEFT, padx=5)

        tb.Button(top, text="Guardar", bootstyle="primary", command=lambda: [actualizar_cantidad(nombre, cantidad_temp.get()), refrescar_tabla(), refrescar_productos_superior(), top.destroy()]).pack(pady=10)

    def cambiar_imagen(nombre):
        top = Toplevel(app)
        top.title(f"Cambiar Imagen - {nombre}")
        top.geometry("300x150")
        top.transient(app)
        top.grab_set()

        def seleccionar():
            archivo = filedialog.askopenfilename(title="Seleccionar imagen", filetypes=[("Imágenes", "*.png *.jpg *.jpeg *.gif")])
            if archivo:
                entry_imagen.delete(0, END)
                entry_imagen.insert(0, archivo)

        tb.Label(top, text="Nueva ruta de imagen:").pack(pady=5)
        entry_imagen = tb.Entry(top)
        entry_imagen.pack(pady=5)
        tb.Button(top, text="Buscar", bootstyle="info", command=seleccionar).pack(pady=5)

        tb.Button(top, text="Guardar", bootstyle="primary", command=lambda: [actualizar_imagen(nombre, entry_imagen.get()), refrescar_productos_superior(), top.destroy()]).pack(pady=10)

    def eliminar_producto_seleccionado():
        seleccionado = inventario_total.focus()
        if not seleccionado:
            messagebox.showwarning("Advertencia", "Debes seleccionar un producto en la lista inferior")
            return
        datos = inventario_total.item(seleccionado, "values")
        if datos and messagebox.askyesno("Confirmar", f"¿Eliminar '{datos[1]}'?"):
            eliminar_producto_db(datos[1])
            refrescar_tabla()
            refrescar_productos_superior()

    def agregar_producto():
        top = Toplevel(app)
        top.title("Agregar Producto")
        top.geometry("300x300")
        top.transient(app)
        top.grab_set()

        tb.Label(top, text="Nombre:").pack(pady=5)
        entry_nombre = tb.Entry(top)
        entry_nombre.pack(pady=5)

        tb.Label(top, text="Cantidad inicial:").pack(pady=5)
        entry_cantidad = tb.Entry(top)
        entry_cantidad.pack(pady=5)

        tb.Label(top, text="Ruta de imagen:").pack(pady=5)
        entry_imagen = tb.Entry(top)
        entry_imagen.pack(pady=5)

        tb.Button(top, text="Buscar", bootstyle="info", command=lambda: seleccionar_archivo(entry_imagen)).pack(pady=5)

        tb.Button(top, text="Guardar", bootstyle="primary", command=lambda: guardar_producto(entry_nombre, entry_cantidad, entry_imagen, top)).pack(pady=10)

    def seleccionar_archivo(entry):
        archivo = filedialog.askopenfilename(title="Seleccionar imagen", filetypes=[("Imágenes", "*.png *.jpg *.jpeg *.gif")])
        if archivo:
            entry.delete(0, END)
            entry.insert(0, archivo)

    def guardar_producto(entry_nombre, entry_cantidad, entry_imagen, top):
        try:
            nombre = entry_nombre.get()
            cantidad = int(entry_cantidad.get())
            imagen = entry_imagen.get()
            if not os.path.exists(imagen):
                raise Exception("La ruta no existe")
            Image.open(imagen)
            agregar_producto_db(nombre, cantidad, imagen)
            refrescar_tabla()
            refrescar_productos_superior()
            top.destroy()
        except:
            messagebox.showerror("Error", "Datos inválidos o imagen incorrecta")

    entry_buscar.bind("<KeyRelease>", lambda e: refrescar_tabla())
    combo_filtros.bind("<<ComboboxSelected>>", lambda e: refrescar_tabla())

    refrescar_tabla()
    refrescar_productos_superior()
    app.mainloop()
